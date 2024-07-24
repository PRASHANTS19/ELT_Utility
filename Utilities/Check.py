import math
import pandas as pd
from mysql.connector import Error
import numpy as np
from openpyxl import load_workbook
from Database.database import DB


def null_check(source_df, target_df, columns=None, writer = None):
    try:
        if source_df is None or target_df is None:
            print("DataFrames are not loaded.")
            return
        if columns:
            source_df[columns] = source_df[columns].replace('', np.nan)
            target_df[columns] = target_df[columns].replace('', np.nan)
        else:
            source_df = source_df.replace('', np.nan)
            target_df = target_df.replace('', np.nan)
        print("Null values comparison:")
        source_null_count = source_df[columns].isnull().sum() if columns else source_df.isnull().sum()
        target_null_count = target_df[columns].isnull().sum() if columns else target_df.isnull().sum()

        print(f"Null count in source table:\n {source_null_count} total: {source_null_count.sum()}")
        print(f"Null count in target table:\n {target_null_count}  total: {target_null_count.sum()}")
        if writer:
            # Write the null count comparison to an Excel sheet
            df_null_comparison = pd.DataFrame({
                'Source Null Count': source_null_count,
                'Target Null Count': target_null_count
            })
            df_null_comparison.to_excel(writer, sheet_name='Null Check')
    except Exception as e:
        print(f"Error during null check: {e}")


def count_check(source_df, target_df, writer=None):
    try:
        source_df_totalrows = len(source_df)
        target_df_totalrows = len(target_df)

        print(f"Total rows in Source table:\n {source_df_totalrows}")
        print(f"Total rows in target table:\n {target_df_totalrows}")
        if writer:
            # Write the count check to an Excel sheet
            df_count_comparison = pd.DataFrame({
                'Source Rows': [source_df_totalrows],
                'Target Rows': [target_df_totalrows]
            })
            df_count_comparison.to_excel(writer, sheet_name='Count Check', index=False)

    except Exception as e:
        print(f"Error during count check: {e}")
    
  
def compare_tables(source_df, target_df, key_column, data_columns, writer=None):
    try:
        # Ensure the specified columns exist in both DataFrames
        columns = [key_column] + data_columns
        for col in columns:
            if col not in source_df.columns or col not in target_df.columns:
                print(f"Column {col} not found in one of the DataFrames.")
                return

        # Fill NaN values with empty strings
        source_df = source_df[columns].fillna('')
        target_df = target_df[columns].fillna('')
        # print("source df is: ", source_df)
        # print("Target df is: ", target_df)

        # Create tuples of the specified columns
        source_tuples = set(map(tuple, source_df.to_records(index=False)))
        target_tuples = set(map(tuple, target_df.to_records(index=False)))

        # print("source tuple is: ", source_tuples)
        # print("Target tuple is: ", target_tuples)

        # Find common rows
        common_rows = target_tuples & source_tuples
        common_count = len(common_rows)

        print(f"Number of common rows: {common_count}")
        if common_count > 0:
            print(f"Common rows: {common_rows}")

        # Find different rows
        source_only_rows = source_tuples - target_tuples
        target_only_rows = target_tuples - source_tuples
        source_only_count = len(source_only_rows)
        target_only_count = len(target_only_rows)

        print(f"Number of rows only in source: {source_only_count}")
        if source_only_count > 0:
            print(f"Rows only in source: {source_only_rows}")
        print(f"Number of rows only in target: {target_only_count}")
        if target_only_count > 0:
            print(f"Rows only in target: {target_only_rows}")

        if writer:
            # Write the table comparison to an Excel sheet
            df_common_rows = pd.DataFrame(list(common_rows), columns=columns)
            df_common_rows.to_excel(writer, sheet_name='Common Rows', index=False)

            df_source_only_rows = pd.DataFrame(list(source_only_rows), columns=columns)
            df_source_only_rows.to_excel(writer, sheet_name='Source Only Rows', index=False)

            df_target_only_rows = pd.DataFrame(list(target_only_rows), columns=columns)
            df_target_only_rows.to_excel(writer, sheet_name='Target Only Rows', index=False)


    except Exception as e:
        print(f"Error during table comparison: {e}")

def ReadData(excel_path: str, sheet_name: str) -> pd.DataFrame:
    try:
        wb = load_workbook(excel_path)
        if sheet_name not in wb.sheetnames:
            raise ValueError(f"Sheet {sheet_name} not found in the Excel file.")
        
        sheet = wb[sheet_name]

        # Read the source type and path from the Excel sheet
        type = sheet['B1'].value.strip()  
        path = sheet['B2'].value.strip() 
        # print(f"Source Type: {type}")
        # print(f"Path: {path}")

        df = None

        # Load the data based on the source type
        if type == 'csv':
            df = pd.read_csv(path)
        elif type == 'excel':
            df = pd.read_excel(path)
        elif type == 'json':
            df = pd.read_json(path)
        elif type == 'database':
            # add db type, postgres
            user = sheet['B5'].value
            password = sheet['B6'].value
            host = sheet['B7'].value
            database_name = sheet['B8'].value 
            table_name = sheet['B9'].value

            db = DB(user, password, host, database_name)
            db.connectDb()
            df = db.readDatabase(table_name)
            db.closeDb()

        else:
            raise ValueError(f"Unsupported source type: {type}")

        return df

    except FileNotFoundError as e:
        print(f"Error: The file at path {path} was not found. {e}")
    except pd.errors.EmptyDataError as e:
        print(f"Error: No data found at path {path}. {e}")
    except ValueError as e:
        print(f"Error: {e}")
    except Exception as e:
        print(f"An unexpected error occurred: {e}")

    return None


def columnCount(source_df, target_df, writer=None):
    try:
        source_column_count = len(source_df.columns)
        target_column_count = len(target_df.columns)

        print(f"Total columns in Source is: {source_column_count}")
        print(f"Total columns in Target is: {target_column_count}")

        if writer:
            # Write the column count comparison to an Excel sheet
            df_column_count_comparison = pd.DataFrame({
                'Source Columns': [source_column_count],
                'Target Columns': [target_column_count]
            })
            df_column_count_comparison.to_excel(writer, sheet_name='Column Count', index=False)
    except Exception as e:
        print(f"Error during column count check: {e}")

def read_excel_data(excel_path, sheet_name):
    try:
        wb = load_workbook(excel_path)
        sheet = wb[sheet_name]
        return sheet
    except FileNotFoundError:
        print(f"Error: The file '{excel_path}' was not found.")
        return None
    except KeyError:
        print(f"Error: The sheet '{sheet_name}' does not exist in the workbook.")
        return None
    except Exception as e:
        print(f"An unexpected error occurred while reading the Excel file: {e}")
        return None
