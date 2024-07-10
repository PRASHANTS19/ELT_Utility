import pandas as pd
from openpyxl import load_workbook
from Utils import Utils


def ReadData(excel_path: str, sheet_name: str) -> pd.DataFrame:
    try:
        wb = load_workbook(excel_path)
        if sheet_name not in wb.sheetnames:
            raise ValueError(f"Sheet {sheet_name} not found in the Excel file.")
        
        sheet = wb[sheet_name]

        # Read the source type and path from the Excel sheet
        type = sheet['B1'].value.strip()  
        path = sheet['B2'].value.strip() 
        # print(f"Source Type: {type}")
        # print(f"Path: {path}")

        df = None

        # Load the data based on the source type
        if type == 'csv':
            df = pd.read_csv(path)
        elif type == 'excel':
            df = pd.read_excel(path)
        elif type == 'json':
            df = pd.read_json(path)
        elif type == 'database':
            user = sheet['B5'].value
            password = sheet['B6'].value
            host = sheet['B7'].value
            database_name = sheet['B8'].value 
            table_name = sheet['B9'].value
            utils = Utils(user, password, host, database_name)
            utils.connectDb()
            df = utils.readDatabase(table_name)
            utils.closeDb()

        else:
            raise ValueError(f"Unsupported source type: {type}")

        return df

    except FileNotFoundError as e:
        print(f"Error: The file at path {path} was not found. {e}")
    except pd.errors.EmptyDataError as e:
        print(f"Error: No data found at path {path}. {e}")
    except ValueError as e:
        print(f"Error: {e}")
    except Exception as e:
        print(f"An unexpected error occurred: {e}")

    return None
